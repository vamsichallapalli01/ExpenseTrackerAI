from django.contrib import messages
from django.core.mail import send_mail
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from datetime import date
from django.db.models import Sum
from django.http import FileResponse, HttpResponse
from django.core.serializers import serialize

from reportlab.pdfgen import canvas
import openpyxl
import json
import pytesseract
import re
from PIL import Image

from rest_framework import generics
from rest_framework.response import Response
from rest_framework.views import APIView

from .utils import create_pdf_report
from .models import (
    Budget,
    Income,
    Expense,
    UserProfile,
    RecurringExpense,
    Receipt,
)
from .forms import (
    RegisterForm,
    IncomeForm,
    ExpenseForm,
    BudgetForm,
    ProfileForm,
    RecurringExpenseForm,
    ReceiptForm,
    RestoreForm,
)
from .serializers import (
    IncomeSerializer,
    ExpenseSerializer,
    BudgetSerializer,
)


def extract_amount(text):
    amounts = re.findall(r"\d+\.\d{2}", text)

    if amounts:
        return max(float(a) for a in amounts)

    return 0


def detect_category(text):
    text = text.lower()

    if "walmart" in text:
        return "Shopping"
    elif "amazon" in text:
        return "Shopping"
    elif "costco" in text:
        return "Shopping"
    elif "shell" in text:
        return "Gas"
    elif "exxon" in text:
        return "Gas"
    elif "chevron" in text:
        return "Gas"
    elif "mcdonald" in text:
        return "Food"
    elif "burger king" in text:
        return "Food"
    elif "subway" in text:
        return "Food"
    elif "cvs" in text:
        return "Healthcare"
    elif "walgreens" in text:
        return "Healthcare"
    else:
        return "Other"


def detect_store(text):
    stores = [
        "Walmart",
        "Amazon",
        "Costco",
        "Shell",
        "Exxon",
        "Chevron",
        "CVS",
        "Walgreens",
        "McDonald's",
        "Burger King",
        "Subway",
    ]

    for store in stores:
        if store.lower() in text.lower():
            return store

    return "Unknown Store"


def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)

        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect("dashboard")

    else:
        form = RegisterForm()

    return render(request, "register.html", {"form": form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)

        if form.is_valid():
            username = form.cleaned_data.get("username")
            password = form.cleaned_data.get("password")

            user = authenticate(
                username=username,
                password=password,
            )

            if user is not None:
                login(request, user)
                return redirect("dashboard")

    else:
        form = AuthenticationForm()

    return render(request, "login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def dashboard(request):
    incomes = Income.objects.filter(user=request.user)
    expenses = Expense.objects.filter(user=request.user)

    current_month = date.today().month
    current_year = date.today().year

    monthly_incomes = incomes.filter(
        date__year=current_year,
        date__month=current_month,
    )

    monthly_expenses_qs = expenses.filter(
        date__year=current_year,
        date__month=current_month,
    )

    total_income = sum(income.amount for income in incomes)
    total_expense = sum(expense.amount for expense in expenses)
    savings = total_income - total_expense

    monthly_income_total = sum(income.amount for income in monthly_incomes)
    monthly_expense_total = sum(expense.amount for expense in monthly_expenses_qs)
    monthly_savings = monthly_income_total - monthly_expense_total

    recent_expenses = expenses.order_by("-date")[:5]

    budget = Budget.objects.filter(user=request.user).first()

    currency_symbol = "$"

    if budget:
        if budget.currency == "USD":
            currency_symbol = "$"
        elif budget.currency == "INR":
            currency_symbol = "₹"
        elif budget.currency == "EUR":
            currency_symbol = "€"
        elif budget.currency == "GBP":
            currency_symbol = "£"

    monthly_budget = 0
    remaining_budget = 0
    budget_percent = 0

    if budget:
        monthly_budget = budget.monthly_budget
        remaining_budget = float(monthly_budget) - float(monthly_expense_total)

        if float(monthly_budget) > 0:
            budget_percent = min(
                round(
                    (float(monthly_expense_total) / float(monthly_budget)) * 100,
                    2,
                ),
                100,
            )

    monthly_expenses = []

    for month in range(1, 13):
        total = Expense.objects.filter(
            user=request.user,
            date__year=current_year,
            date__month=month,
        ).aggregate(
            Sum("amount")
        )["amount__sum"] or 0

        monthly_expenses.append(float(total))

    prediction = 0

    non_zero_months = [
        amount for amount in monthly_expenses if amount > 0
    ]

    if len(non_zero_months) >= 3:
        last_month = non_zero_months[-1]
        previous_month = non_zero_months[-2]
        growth = last_month - previous_month
        prediction = last_month + growth

    elif len(non_zero_months) > 0:
        prediction = sum(non_zero_months) / len(non_zero_months)

    prediction = max(0, prediction)

    if monthly_budget > 0 and prediction > float(monthly_budget):
        send_mail(
            "Budget Warning",
            "Predicted spending exceeds budget.",
            None,
            [request.user.email],
            fail_silently=True,
        )

    predicted_savings = float(monthly_income_total) - float(prediction)
    monthly_saving_forecast = max(0, predicted_savings)

    savings_rate = 0

    if total_income > 0:
        savings_rate = (savings / total_income) * 100

    recommendations = []

    food_total = sum(
        expense.amount
        for expense in Expense.objects.filter(
            user=request.user,
            category="Food",
        )
    )

    food_percent = 0

    if total_expense > 0:
        food_percent = (food_total / total_expense) * 100

    if food_percent > 30:
        recommendations.append(
            "🍔 Food spending exceeds 30% of expenses."
        )

    shopping_total = sum(
        expense.amount
        for expense in Expense.objects.filter(
            user=request.user,
            category="Shopping",
        )
    )

    shopping_percent = 0

    if total_expense > 0:
        shopping_percent = (shopping_total / total_expense) * 100

    if shopping_percent > 20:
        recommendations.append(
            "🛍 Shopping expenses are high this month."
        )

    savings_percent = 0

    if total_income > 0:
        savings_percent = (savings / total_income) * 100

    if savings_percent < 20:
        recommendations.append(
            "💰 Try saving at least 20% of income."
        )

    if monthly_budget > 0 and prediction > float(monthly_budget):
        recommendations.append(
            "⚠ Predicted spending exceeds budget."
        )

    if savings_rate >= 25:
        recommendations.append(
            "🎉 Excellent savings discipline."
        )

    category_totals = {}

    for expense in Expense.objects.filter(user=request.user):
        category = expense.category

        if category not in category_totals:
            category_totals[category] = 0

        category_totals[category] += float(expense.amount)

    if category_totals:
        top_category_recommendation = max(
            category_totals,
            key=category_totals.get,
        )

        recommendations.append(
            f"📊 Your highest spending category is {top_category_recommendation}."
        )

        highest_amount = max(category_totals.values())

        potential_savings = round(
            highest_amount * 0.10,
            2,
        )

        recommendations.append(
            f"💡 Reducing your largest category by 10% could save {currency_symbol}{potential_savings}."
        )

    total_transactions = expenses.count()

    savings_rate = 0

    if monthly_income_total > 0:
        savings_rate = round(
            (float(savings) / float(monthly_income_total)) * 100,
            2,
        )

    financial_score = 0

    if savings_rate >= 30:
        financial_score += 40
    elif savings_rate >= 20:
        financial_score += 30
    elif savings_rate >= 10:
        financial_score += 20
    else:
        financial_score += 10

    if monthly_budget > 0:
        budget_usage = (
            float(monthly_expense_total) / float(monthly_budget)
        ) * 100
    else:
        budget_usage = 100

    if budget_usage <= 70:
        financial_score += 30
    elif budget_usage <= 90:
        financial_score += 20
    else:
        financial_score += 10

    if monthly_budget > 0:
        if prediction <= float(monthly_budget):
            financial_score += 20
        else:
            financial_score += 5

    if total_transactions <= 50:
        financial_score += 10
    else:
        financial_score += 5

    if financial_score >= 80:
        financial_status = "Excellent"
    elif financial_score >= 60:
        financial_status = "Good"
    elif financial_score >= 40:
        financial_status = "Average"
    else:
        financial_status = "Poor"

    if financial_status == "Excellent":
        health_message = "🎉 Your finances are very healthy."
    elif financial_status == "Good":
        health_message = "✅ You are managing money well."
    elif financial_status == "Average":
        health_message = "⚠ Consider increasing savings."
    else:
        health_message = "🚨 Immediate financial improvement needed."

    current_savings = float(savings)

    savings_1_month = current_savings + monthly_saving_forecast
    savings_3_months = current_savings + (monthly_saving_forecast * 3)
    savings_6_months = current_savings + (monthly_saving_forecast * 6)
    savings_12_months = current_savings + (monthly_saving_forecast * 12)

    forecast_labels = [
        "Now",
        "1 Month",
        "3 Months",
        "6 Months",
        "12 Months",
    ]

    forecast_values = [
        current_savings,
        savings_1_month,
        savings_3_months,
        savings_6_months,
        savings_12_months,
    ]

    forecast_labels = json.dumps(forecast_labels)
    forecast_values = json.dumps(forecast_values)

    if savings_12_months >= 10000:
        forecast_message = "🎉 You may save over $10,000 within a year."
    elif savings_12_months >= 5000:
        forecast_message = "✅ You are on track for strong savings growth."
    else:
        forecast_message = "⚠ Increase savings to improve future wealth."

    budget_difference = 0

    if float(monthly_budget) > 0:
        budget_difference = float(monthly_budget) - prediction

    overrun_alert = ""

    if float(monthly_budget) > 0:
        if prediction > float(monthly_budget):
            overrun_alert = (
                f"You may exceed your budget by "
                f"{currency_symbol}{abs(round(budget_difference, 2))}"
            )
        else:
            overrun_alert = (
                f"You are expected to stay under budget by "
                f"{currency_symbol}{round(budget_difference, 2)}"
            )

    budget_score = 100

    if float(monthly_budget) > 0:
        budget_score = max(
            0,
            round(
                (
                    float(monthly_budget) - prediction
                )
                / float(monthly_budget)
                * 100
            ),
        )

    if budget_score >= 40:
        budget_status = "Excellent"
    elif budget_score >= 20:
        budget_status = "Good"
    elif budget_score >= 0:
        budget_status = "Warning"
    else:
        budget_status = "Critical"

    risk_message = ""

    if float(monthly_budget) > 0:
        if prediction > float(monthly_budget):
            risk_message = "⚠ Predicted spending exceeds budget"
        else:
            risk_message = "✅ Budget is likely safe"

    category_totals = {}

    for expense in expenses:
        if expense.category not in category_totals:
            category_totals[expense.category] = 0

        category_totals[expense.category] += expense.amount

    top_category = "No Data"

    if category_totals:
        top_category = max(category_totals, key=category_totals.get)

    largest_expense = expenses.order_by("-amount").first()
    largest_expense_amount = 0

    if largest_expense:
        largest_expense_amount = largest_expense.amount

    average_expense = 0

    if expenses.exists():
        average_expense = round(float(total_expense) / expenses.count(), 2)

    expense_categories = []
    expense_amounts = []

    for category, _ in Expense.CATEGORY_CHOICES:
        category_total = expenses.filter(
            category=category
        ).aggregate(
            Sum("amount")
        )["amount__sum"] or 0

        if category_total > 0:
            expense_categories.append(category)
            expense_amounts.append(float(category_total))

    monthly_labels = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]

    monthly_totals = monthly_expenses.copy()

    monthly_labels.append("Prediction")
    monthly_totals.append(round(prediction, 2))

    notifications = []

    if budget_percent >= 90:
        notifications.append("⚠ Budget usage is above 90%.")

    if remaining_budget < 0:
        notifications.append("⚠ You exceeded your monthly budget.")

    if savings < 0:
        notifications.append("⚠ Expenses exceed income.")

    if total_income == 0:
        notifications.append("⚠ No income recorded.")

    if RecurringExpense.objects.filter(user=request.user).exists():
        notifications.append("ℹ Recurring expenses are active.")

    insights = []

    if top_category != "No Data":
        insights.append(
            f"Your largest spending category is {top_category}."
        )

    if total_income > 0:
        savings_percent = round(
            (float(savings) / float(total_income)) * 100,
            2,
        )

        insights.append(
            f"You saved {savings_percent}% of your income."
        )

    if remaining_budget > 0:
        insights.append(
            f"You still have {currency_symbol}{round(remaining_budget, 2)} remaining in your budget."
        )

    context = {
        "total_income": total_income,
        "total_expense": total_expense,
        "savings": savings,
        "currency_symbol": currency_symbol,
        "recent_expenses": recent_expenses,
        "monthly_budget": monthly_budget,
        "remaining_budget": remaining_budget,
        "budget_percent": budget_percent,
        "monthly_income_total": monthly_income_total,
        "monthly_expense_total": monthly_expense_total,
        "monthly_savings": monthly_savings,
        "top_category": top_category,
        "largest_expense_amount": largest_expense_amount,
        "average_expense": average_expense,
        "total_transactions": total_transactions,
        "expense_categories": expense_categories,
        "expense_amounts": expense_amounts,
        "monthly_labels": monthly_labels,
        "monthly_totals": monthly_totals,
        "prediction": round(prediction, 2),
        "predicted_savings": round(predicted_savings, 2),
        "budget_difference": round(budget_difference, 2),
        "overrun_alert": overrun_alert,
        "budget_score": budget_score,
        "budget_status": budget_status,
        "risk_message": risk_message,
        "insights": insights,
        "notifications": notifications,
        "savings_1_month": round(savings_1_month, 2),
        "savings_3_months": round(savings_3_months, 2),
        "savings_6_months": round(savings_6_months, 2),
        "savings_12_months": round(savings_12_months, 2),
        "forecast_labels": forecast_labels,
        "forecast_values": forecast_values,
        "forecast_message": forecast_message,
        "financial_score": financial_score,
        "financial_status": financial_status,
        "health_message": health_message,
        "savings_rate": savings_rate,
        "recommendations": recommendations,
    }

    return render(request, "dashboard.html", context)


@login_required
def add_income(request):
    if request.method == "POST":
        form = IncomeForm(request.POST)

        if form.is_valid():
            income = form.save(commit=False)
            income.user = request.user
            income.save()
            return redirect("dashboard")

    else:
        form = IncomeForm()

    return render(request, "add_income.html", {"form": form})


@login_required
def add_expense(request):
    if request.method == "POST":
        form = ExpenseForm(request.POST)

        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            return redirect("dashboard")

    else:
        form = ExpenseForm()

    return render(request, "add_expense.html", {"form": form})


@login_required
def income_list(request):
    incomes = Income.objects.filter(
        user=request.user
    )

    return render(
        request,
        "income_list.html",
        {
            "incomes": incomes,
        },
    )


@login_required
def set_budget(request):
    budget = Budget.objects.filter(user=request.user).first()

    if request.method == "POST":
        form = BudgetForm(request.POST, instance=budget)

        if form.is_valid():
            budget = form.save(commit=False)
            budget.user = request.user
            budget.save()
            return redirect("dashboard")

    else:
        form = BudgetForm(instance=budget)

    return render(request, "budget.html", {"form": form})


@login_required
def reports(request):
    incomes = Income.objects.filter(
        user=request.user
    ).order_by("-date")

    expenses = Expense.objects.filter(
        user=request.user
    ).order_by("-date")

    category = request.GET.get("category")
    month = request.GET.get("month")

    if category:
        expenses = expenses.filter(
            category=category
        )

    if month:
        expenses = expenses.filter(
            date__month=month
        )

    context = {
        "incomes": incomes,
        "expenses": expenses,
        "categories": Expense.CATEGORY_CHOICES,
    }

    return render(
        request,
        "reports.html",
        context,
    )


@login_required
def export_pdf(request):
    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="expenses.pdf"'
    )

    p = canvas.Canvas(response)

    p.setFont(
        "Helvetica-Bold",
        16,
    )

    p.drawString(
        50,
        800,
        "Expense Tracker Report",
    )

    expenses = Expense.objects.filter(
        user=request.user
    )

    y = 760

    for expense in expenses:
        p.drawString(
            50,
            y,
            f"{expense.date} | {expense.category} | ${expense.amount}",
        )

        y -= 20

    p.save()

    return response


@login_required
def export_excel(request):
    workbook = openpyxl.Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    incomes = Income.objects.filter(
        user=request.user
    )

    expenses = Expense.objects.filter(
        user=request.user
    )

    total_income = sum(
        income.amount
        for income in incomes
    )

    total_expense = sum(
        expense.amount
        for expense in expenses
    )

    savings = total_income - total_expense

    summary_sheet.append(
        ["Expense Tracker Summary"]
    )

    summary_sheet.append([])

    summary_sheet.append(
        ["Total Income", float(total_income)]
    )

    summary_sheet.append(
        ["Total Expense", float(total_expense)]
    )

    summary_sheet.append(
        ["Savings", float(savings)]
    )

    income_sheet = workbook.create_sheet(
        "Income"
    )

    income_sheet.append(
        ["Source", "Amount", "Date"]
    )

    for income in incomes:
        income_sheet.append(
            [
                income.source,
                float(income.amount),
                str(income.date),
            ]
        )

    expense_sheet = workbook.create_sheet(
        "Expenses"
    )

    expense_sheet.append(
        [
            "Category",
            "Amount",
            "Date",
            "Description",
        ]
    )

    for expense in expenses:
        expense_sheet.append(
            [
                expense.category,
                float(expense.amount),
                str(expense.date),
                expense.description,
            ]
        )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="finance_report.xlsx"'
    )

    workbook.save(response)

    return response


@login_required
def profile(request):
    profile, created = UserProfile.objects.get_or_create(
        user=request.user
    )

    if request.method == "POST":
        form = ProfileForm(
            request.POST,
            request.FILES,
            instance=profile,
        )

        if form.is_valid():
            form.save()
            return redirect("profile")

    else:
        form = ProfileForm(instance=profile)

    return render(
        request,
        "profile.html",
        {
            "form": form,
            "profile": profile,
        },
    )


@login_required
def recurring_expenses(request):
    expenses = RecurringExpense.objects.filter(
        user=request.user
    )

    return render(
        request,
        "recurring_expenses.html",
        {
            "expenses": expenses,
        },
    )


@login_required
def add_recurring_expense(request):
    if request.method == "POST":
        form = RecurringExpenseForm(
            request.POST
        )

        if form.is_valid():
            expense = form.save(
                commit=False
            )

            expense.user = request.user
            expense.save()

            return redirect(
                "recurring_expenses"
            )

    else:
        form = RecurringExpenseForm()

    return render(
        request,
        "add_recurring_expense.html",
        {
            "form": form,
        },
    )


@login_required
def upload_receipt(request):
    extracted_text = ""
    category = None
    amount = None
    store = None

    if request.method == "POST":
        form = ReceiptForm(request.POST, request.FILES)

        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.user = request.user
            receipt.save()

            try:
                image = Image.open(receipt.image.path)
                extracted_text = pytesseract.image_to_string(image)

                category = detect_category(extracted_text)
                amount = extract_amount(extracted_text)
                store = detect_store(extracted_text)

                if amount > 0:
                    Expense.objects.create(
                        user=request.user,
                        category=category,
                        amount=amount,
                        description=f"Receipt from {store}",
                        date=date.today(),
                    )

                    messages.success(
                        request,
                        "Receipt processed and expense added successfully.",
                    )

                    return redirect("dashboard")

                else:
                    messages.warning(
                        request,
                        "Receipt uploaded, but amount was not detected.",
                    )

            except Exception as e:
                messages.error(
                    request,
                    f"Receipt scanning failed: {e}",
                )

    else:
        form = ReceiptForm()

    return render(
        request,
        "upload_receipt.html",
        {
            "form": form,
            "text": extracted_text,
            "category": category,
            "amount": amount,
            "store": store,
        },
    )


@login_required
def send_report(request):
    total_income = sum(
        income.amount
        for income in Income.objects.filter(user=request.user)
    )

    total_expense = sum(
        expense.amount
        for expense in Expense.objects.filter(user=request.user)
    )

    savings = total_income - total_expense

    financial_score = 80 if savings > 0 else 50

    budget_status = "🟢 Healthy" if savings > 0 else "🔴 Over Budget"

    recommendations = []

    if savings > 0:
        recommendations.append(
            "You are saving money. Continue tracking your expenses regularly."
        )
    else:
        recommendations.append(
            "Your expenses are higher than your income. Try reducing unnecessary spending."
        )

    if financial_score >= 80:
        recommendations.append(
            "Your financial health is good. Keep maintaining your current habits."
        )
    else:
        recommendations.append(
            "Focus on improving your savings and controlling monthly expenses."
        )

    recommendation_text = "\n".join(recommendations)

    currency_symbol = "$"

    message = f"""

📊 Monthly Financial Report

Total Income:
{currency_symbol}{total_income}

Total Expenses:
{currency_symbol}{total_expense}

Savings:
{currency_symbol}{savings}

Financial Health:
{financial_score}/100

Budget Status:
{budget_status}

AI Recommendations

{recommendation_text}

"""

    send_mail(
        "Expense Tracker Report",
        message,
        None,
        [request.user.email],
        fail_silently=False,
    )

    return redirect("dashboard")


class IncomeListAPI(generics.ListAPIView):
    serializer_class = IncomeSerializer

    def get_queryset(self):
        return Income.objects.filter(
            user=self.request.user
        )


class ExpenseListAPI(generics.ListAPIView):
    serializer_class = ExpenseSerializer

    def get_queryset(self):
        return Expense.objects.filter(
            user=self.request.user
        )


class BudgetAPI(generics.ListAPIView):
    serializer_class = BudgetSerializer

    def get_queryset(self):
        return Budget.objects.filter(
            user=self.request.user
        )


class DashboardAPI(APIView):
    def get(self, request):
        incomes = Income.objects.filter(
            user=request.user
        )

        expenses = Expense.objects.filter(
            user=request.user
        )

        total_income = sum(
            i.amount
            for i in incomes
        )

        total_expense = sum(
            e.amount
            for e in expenses
        )

        savings = total_income - total_expense

        return Response(
            {
                "income": total_income,
                "expense": total_expense,
                "savings": savings,
            }
        )


@login_required
def edit_income(request, pk):
    income = get_object_or_404(
        Income,
        pk=pk,
        user=request.user,
    )

    if request.method == "POST":
        form = IncomeForm(
            request.POST,
            instance=income,
        )

        if form.is_valid():
            form.save()
            return redirect("dashboard")

    else:
        form = IncomeForm(instance=income)

    return render(
        request,
        "edit_income.html",
        {
            "form": form,
        },
    )


@login_required
def delete_income(request, pk):
    income = get_object_or_404(
        Income,
        pk=pk,
        user=request.user,
    )

    income.delete()

    return redirect("dashboard")


@login_required
def edit_expense(request, pk):
    expense = get_object_or_404(
        Expense,
        pk=pk,
        user=request.user,
    )

    if request.method == "POST":
        form = ExpenseForm(
            request.POST,
            instance=expense,
        )

        if form.is_valid():
            form.save()
            return redirect("dashboard")

    else:
        form = ExpenseForm(instance=expense)

    return render(
        request,
        "edit_expense.html",
        {
            "form": form,
        },
    )


@login_required
def delete_expense(request, pk):
    expense = get_object_or_404(
        Expense,
        pk=pk,
        user=request.user,
    )

    expense.delete()

    return redirect("dashboard")


@login_required
def expense_list(request):
    expenses = Expense.objects.filter(
        user=request.user
    )

    category = request.GET.get("category")
    expense_date = request.GET.get("date")
    amount = request.GET.get("amount")
    month = request.GET.get("month")
    year = request.GET.get("year")
    sort = request.GET.get("sort")

    if category:
        expenses = expenses.filter(
            category__icontains=category
        )

    if expense_date:
        expenses = expenses.filter(
            date=expense_date
        )

    if amount:
        expenses = expenses.filter(
            amount=amount
        )

    if month:
        expenses = expenses.filter(
            date__month=month
        )

    if year:
        expenses = expenses.filter(
            date__year=year
        )

    if sort == "high":
        expenses = expenses.order_by(
            "-amount"
        )

    elif sort == "low":
        expenses = expenses.order_by(
            "amount"
        )

    else:
        expenses = expenses.order_by(
            "-date"
        )

    filtered_total = sum(
        expense.amount
        for expense in expenses
    )

    currency_symbol = "$"

    if hasattr(request.user, "profile"):
        currency_symbol = request.user.profile.currency_symbol

    return render(
        request,
        "expense_list.html",
        {
            "expenses": expenses,
            "filtered_total": filtered_total,
            "currency_symbol": currency_symbol,
        },
    )


@login_required
def backup_data(request):
    incomes = Income.objects.filter(
        user=request.user
    )

    expenses = Expense.objects.filter(
        user=request.user
    )

    budgets = Budget.objects.filter(
        user=request.user
    )

    data = {
        "income": json.loads(
            serialize(
                "json",
                incomes,
            )
        ),
        "expense": json.loads(
            serialize(
                "json",
                expenses,
            )
        ),
        "budget": json.loads(
            serialize(
                "json",
                budgets,
            )
        ),
    }

    response = HttpResponse(
        json.dumps(
            data,
            indent=4,
        ),
        content_type="application/json",
    )

    response["Content-Disposition"] = (
        'attachment; filename="backup.json"'
    )

    return response


@login_required
def restore_data(request):
    if request.method == "POST":
        form = RestoreForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            uploaded_file = request.FILES["backup_file"]

            data = json.load(uploaded_file)

            Income.objects.filter(
                user=request.user
            ).delete()

            Expense.objects.filter(
                user=request.user
            ).delete()

            Budget.objects.filter(
                user=request.user
            ).delete()

            for item in data["income"]:
                Income.objects.create(
                    user=request.user,
                    source=item["fields"]["source"],
                    amount=item["fields"]["amount"],
                    date=item["fields"]["date"],
                )

            for item in data["expense"]:
                Expense.objects.create(
                    user=request.user,
                    category=item["fields"]["category"],
                    amount=item["fields"]["amount"],
                    date=item["fields"]["date"],
                    description=item["fields"]["description"],
                )

            for item in data["budget"]:
                Budget.objects.create(
                    user=request.user,
                    monthly_budget=item["fields"]["monthly_budget"],
                )

            messages.success(
                request,
                "Backup restored successfully.",
            )

            return redirect("dashboard")

    else:
        form = RestoreForm()

    return render(
        request,
        "restore.html",
        {
            "form": form,
        },
    )